"""Обработка команды /report для генерации отчётов."""
import csv
import io
import logging

from datetime import datetime, date

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from aiogram import Router, F
from aiogram.types import Message, BufferedInputFile, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.filters import Command

from db import (
    get_orders_for_report,
    get_orders_by_status,
    get_orders_for_date,
    get_active_orders_for_date,
    get_daily_report_enabled,
    set_daily_report_enabled,
    get_report_chat_id,
    set_report_chat_id,
    bulk_update_order_status,
    is_user_authorized,
)
from keyboards import (
    BUTTON_REPORT,
    BUTTON_NEW,
    BUTTON_IN_PROGRESS,
    BUTTON_PAID,
    BUTTON_DELIVERED,
    BUTTON_CANCELED,
    BUTTON_SEARCH,
    BUTTON_DAILY_REPORT_TOGGLE_ON,
    BUTTON_DAILY_REPORT_TOGGLE_OFF,
)

logger = logging.getLogger(__name__)
router = Router()

STATUS_BY_BUTTON = {
    BUTTON_NEW: "new",
    BUTTON_IN_PROGRESS: "in_progress",
    BUTTON_PAID: "paid",
    BUTTON_DELIVERED: "delivery",
    BUTTON_CANCELED: "canceled",
}

STATUS_DISPLAY = {
    "new": "🆕 Новый",
    "in_progress": "📦 В работе",
    "delivery": "🚚 Доставка",
    "paid": "✅ Оплачен",
    "canceled": "❌ Отказ",
}

# Регистрируем TTF-шрифт с поддержкой кириллицы.
# Файл adomino.ttf должен лежать в корне проекта (там, где запускается бот).
try:
    pdfmetrics.registerFont(TTFont("Adomino", "adomino.ttf"))
    logger.info("TTF-шрифт 'Adomino' успешно зарегистрирован")
except Exception as e:
    logger.error(f"Не удалось зарегистрировать шрифт adomino.ttf: {e}")


@router.message(F.text == BUTTON_REPORT)
async def report_button(message: Message) -> None:
    """Обработка кнопки '📊 Отчёт' для генерации PDF-отчёта."""
    if not await is_user_authorized(message.from_user.id):
        await message.reply("🔐 Доступ запрещён. Введите /start и PIN-код.")
        return
    logger.info(f"Обработка кнопки '📊 Отчёт' от пользователя {message.from_user.id}")
    await cmd_report_pdf(message)


def _format_order_price(price: str) -> str:
    """Форматирует цену."""
    return price if price else ""


def _format_order_line(order: dict) -> str:
    """Форматирует строку заказа для отображения в новом формате."""
    from datetime import datetime, timedelta
    
    order_id = order.get("id", "")
    status = order.get("status", "new")
    
    # Импортируем STATUS_OPTIONS из handlers.orders
    STATUS_OPTIONS = {
        "new": "🆕",
        "in_progress": "📦",
        "delivery": "🚚",
        "paid": "✅",
        "canceled": "❌",
    }
    status_icon = STATUS_OPTIONS.get(status, "🆕")
    
    model = order.get("model", "") or ""
    price = _format_order_price(order.get("price", ""))
    address = order.get("address", "") or ""
    phone = order.get("phone", "") or ""
    comment = order.get("comment", "") or ""
    manager_name = order.get("manager_name", "") or ""
    
    # Форматируем дату/время
    created_at = order.get("created_at", "")
    date_str = ""
    if created_at:
        try:
            if "T" in created_at:
                dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
            else:
                dt = datetime.fromisoformat(created_at)
            
            now = datetime.now()
            today = now.date()
            order_date = dt.date()
            
            if order_date == today:
                date_str = f"сегодня {dt.strftime('%H:%M')}"
            elif order_date == today - timedelta(days=1):
                date_str = f"вчера {dt.strftime('%H:%M')}"
            else:
                date_str = dt.strftime("%d.%m %H:%M")
        except Exception:
            if "T" in created_at:
                date_str = created_at.split("T")[0]
            else:
                date_str = created_at
    
    # Обрезаем длинные поля
    if len(comment) > 20:
        comment = comment[:20] + "..."
    if len(model) > 30:
        model = model[:30] + "..."
    
    parts = [
        f"#{order_id}",
        status_icon,
        model,
        price,
        address,
        phone,
        comment,
        date_str,
        manager_name,
    ]
    
    # Убираем пустые части
    parts = [p for p in parts if p]
    
    return " • ".join(parts)


@router.message(F.text.in_(STATUS_BY_BUTTON.keys()))
async def handle_status_button(message: Message) -> None:
    """Обработка кнопок со статусами для показа списка заказов (последние 10)."""
    if not await is_user_authorized(message.from_user.id):
        await message.reply("🔐 Доступ запрещён. Введите /start и PIN-код.")
        return
    logger.info(f"Обработка кнопки статуса '{message.text}' от пользователя {message.from_user.id}")
    try:
        status = STATUS_BY_BUTTON[message.text]
        orders = await get_orders_by_status(status, limit=10)

        if not orders:
            await message.reply("📭 Заказов с таким статусом пока нет")
            return

        lines = []
        if len(orders) >= 10:
            lines.append("Показаны последние 10 записей:")
        else:
            lines.append(f"Показаны последние {len(orders)} записей:")

        for order in orders:
            lines.append(_format_order_line(order))

        text = "\n".join(lines)

        from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📊 Сформировать отчёт", callback_data=f"report_status:{status}")],
            [InlineKeyboardButton(text="🔄 Изменить статус всех", callback_data=f"bulk_status_menu:{status}")],
        ])

        await message.reply(text, reply_markup=keyboard)

        logger.info(
            f"Список заказов со статусом {status} отправлен пользователю {message.from_user.id}, показано: {len(orders)}"
        )

    except Exception as e:
        logger.error(f"Ошибка при получении заказов по статусу: {e}", exc_info=True)
        await message.reply("⚠️ Произошла ошибка при получении списка заказов")


@router.callback_query(F.data.startswith("report_status:"))
async def handle_report_status_callback(callback: CallbackQuery) -> None:
    """Генерация отчёта по статусу."""
    try:
        status = callback.data.split(":", 1)[1]
        orders = await get_orders_by_status(status)

        if not orders:
            await callback.answer("Заказов с таким статусом нет", show_alert=True)
            return

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заказы"

        headers = [
            "ID", "Дата", "Менеджер", "Статус", "Заказ", "Цена",
            "Адрес", "Телефон", "Клиент", "Комментарий",
        ]

        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_num, order in enumerate(orders, 2):
            created_at = order.get("created_at", "") or ""
            if "T" in created_at:
                created_at = created_at.split("T")[0]

            values = [
                order.get("id", ""), created_at, order.get("manager_name", ""),
                order.get("status", ""), order.get("model", ""), order.get("price", ""),
                order.get("address", ""), order.get("phone", ""),
                order.get("customer_name", ""), order.get("comment", ""),
            ]

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_bytes = buffer.read()
        buffer.close()

        from aiogram.types import BufferedInputFile
        xlsx_file = BufferedInputFile(xlsx_bytes, filename=f"report_{status}.xlsx")

        await callback.message.answer_document(
            document=xlsx_file,
            caption=f"📊 Отчёт по статусу: {STATUS_DISPLAY.get(status, status)}",
        )
        await callback.answer("✅ Отчёт сформирован")

    except Exception as e:
        logger.error(f"Ошибка при генерации отчёта по статусу: {e}", exc_info=True)
        await callback.answer("⚠️ Ошибка при генерации отчёта", show_alert=True)


@router.callback_query(F.data.startswith("bulk_status_menu:"))
async def handle_bulk_status_menu(callback: CallbackQuery) -> None:
    """Показывает меню для массового изменения статуса."""
    if not await is_user_authorized(callback.from_user.id):
        await callback.answer("🔐 Доступ запрещён. Введите /start и PIN-код.", show_alert=True)
        return
    try:
        old_status = callback.data.split(":", 1)[1]
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="🆕 Новый", callback_data=f"bulk_status:{old_status}:new"),
                InlineKeyboardButton(text="📦 В работе", callback_data=f"bulk_status:{old_status}:in_progress"),
            ],
            [
                InlineKeyboardButton(text="🚚 Доставка", callback_data=f"bulk_status:{old_status}:delivery"),
                InlineKeyboardButton(text="✅ Оплачен", callback_data=f"bulk_status:{old_status}:paid"),
            ],
            [
                InlineKeyboardButton(text="❌ Отказ", callback_data=f"bulk_status:{old_status}:canceled"),
            ],
        ])
        await callback.message.edit_text(
            "Хотите изменить статус всех заказов?\nВыберите новый статус:",
            reply_markup=keyboard,
        )
        await callback.answer()
    except Exception as e:
        logger.error(f"Ошибка в bulk_status_menu: {e}", exc_info=True)
        await callback.answer("⚠️ Ошибка", show_alert=True)


@router.callback_query(F.data.startswith("bulk_status:"))
async def handle_bulk_status_callback(callback: CallbackQuery) -> None:
    """Массовое изменение статуса заказов."""
    try:
        parts = callback.data.split(":")
        if len(parts) != 3:
            await callback.answer("⚠️ Ошибка формата", show_alert=True)
            return

        old_status = parts[1]
        new_status = parts[2]

        orders = await get_orders_by_status(old_status)
        if not orders:
            await callback.answer("Нет заказов для изменения", show_alert=True)
            return

        order_ids = [o["id"] for o in orders]
        updated_count = await bulk_update_order_status(order_ids, new_status)

        new_status_display = STATUS_DISPLAY.get(new_status, new_status)
        await callback.message.edit_text(
            f"✅ Статус {updated_count} заказов обновлён на {new_status_display}",
        )
        await callback.answer(f"✅ Обновлено: {updated_count}")

        logger.info(f"Массовое изменение статуса: {len(order_ids)} заказов с {old_status} на {new_status}")

    except Exception as e:
        logger.error(f"Ошибка при массовом изменении статуса: {e}", exc_info=True)
        await callback.answer("⚠️ Произошла ошибка", show_alert=True)


@router.message(Command("report"))
async def cmd_report(message: Message) -> None:
    """Генерирует и отправляет CSV-отчёт по заказам."""
    try:
        orders = await get_orders_for_report()

        if not orders:
            await message.reply("📋 Заказов пока нет")
            return

        output = io.StringIO()

        fieldnames = [
            "id",
            "created_at",
            "manager_name",
            "status",
            "model",
            "price",
            "address",
            "phone",
            "customer_name",
            "comment",
        ]

        writer = csv.DictWriter(output, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()

        for order in orders:
            row = {key: "" for key in fieldnames}
            for key in fieldnames:
                value = order.get(key)
                row[key] = "" if value is None else value
            writer.writerow(row)

        csv_content = output.getvalue()
        output.close()

        csv_bytes = csv_content.encode("utf-8-sig")
        csv_file = BufferedInputFile(csv_bytes, filename="report.csv")

        await message.reply_document(
            document=csv_file,
            caption="📊 Отчёт по заказам (CSV)",
        )

        logger.info(
            f"CSV-отчёт отправлен пользователю {message.from_user.id}, записей: {len(orders)}"
        )

    except Exception as e:
        logger.error(f"Ошибка при генерации CSV-отчёта: {e}", exc_info=True)
        await message.reply("⚠️ Произошла ошибка при генерации отчёта")


@router.message(Command("report_pdf"))
async def cmd_report_pdf(message: Message) -> None:
    """Генерирует и отправляет PDF-отчёт по заказам с кириллицей."""
    if not await is_user_authorized(message.from_user.id):
        await message.reply("🔐 Доступ запрещён. Введите /start и PIN-код.")
        return
    try:
        orders = await get_orders_for_report()

        if not orders:
            await message.reply("📋 Заказов пока нет")
            return

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()

        # Меняем шрифт во всех базовых стилях на Adomino
        for style in styles.byName.values():
            style.fontName = "Adomino"

        story = []

        title = Paragraph("Отчёт по заказам", styles["Title"])
        story.append(title)
        story.append(Spacer(1, 10 * mm))

        # Заголовки таблицы
        data = [
            ["ID", "Дата", "Менеджер", "Статус", "Модель", "Клиент", "Телефон", "Адрес"]
        ]

        for order in orders:
            order_id = str(order.get("id", ""))
            created_at = order.get("created_at", "") or ""
            if "T" in created_at:
                created_at = created_at.split("T")[0]

            manager_name = order.get("manager_name", "") or ""
            status_raw = order.get("status", "") or ""
            status_display = STATUS_DISPLAY.get(status_raw, status_raw)
            model = order.get("model", "") or ""
            customer_name = order.get("customer_name", "") or "Без имени"
            phone = order.get("phone", "") or ""
            address = order.get("address", "") or ""

            data.append(
                [
                    order_id,
                    created_at,
                    manager_name,
                    status_display,
                    model,
                    customer_name,
                    phone,
                    address,
                ]
            )

        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "Adomino"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]
            )
        )

        story.append(table)
        doc.build(story)

        pdf_bytes = buffer.getvalue()
        buffer.close()

        pdf_file = BufferedInputFile(pdf_bytes, filename="report.pdf")

        await message.reply_document(
            document=pdf_file,
            caption="📊 Отчёт по заказам (PDF)",
        )

        logger.info(
            f"PDF-отчёт отправлен пользователю {message.from_user.id}, записей: {len(orders)}"
        )

    except Exception as e:
        logger.error(f"Ошибка при генерации PDF-отчёта: {e}", exc_info=True)
        await message.reply("⚠️ Произошла ошибка при генерации PDF-отчёта")


@router.message(Command("report_xlsx"))
async def cmd_report_xlsx(message: Message) -> None:
    """Генерирует и отправляет Excel-отчёт по заказам."""
    if not await is_user_authorized(message.from_user.id):
        await message.reply("🔐 Доступ запрещён. Введите /start и PIN-код.")
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        orders = await get_orders_for_report()

        if not orders:
            await message.reply("📋 Заказов пока нет")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заказы"

        headers = [
            "ID",
            "Дата",
            "Менеджер",
            "Статус",
            "Заказ",
            "Цена",
            "Адрес",
            "Телефон",
            "Клиент",
            "Комментарий",
        ]

        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        header_font = Font(bold=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Данные
        for row_num, order in enumerate(orders, 2):
            created_at = order.get("created_at", "") or ""
            if "T" in created_at:
                created_at = created_at.split("T")[0]

            values = [
                order.get("id", ""),
                created_at,
                order.get("manager_name", ""),
                order.get("status", ""),
                order.get("model", ""),
                order.get("price", ""),
                order.get("address", ""),
                order.get("phone", ""),
                order.get("customer_name", ""),
                order.get("comment", ""),
            ]

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border

        # Автоширина
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_bytes = buffer.read()
        buffer.close()

        xlsx_file = BufferedInputFile(xlsx_bytes, filename="report.xlsx")

        await message.reply_document(
            document=xlsx_file,
            caption="📊 Отчёт по заказам (Excel)",
        )

        logger.info(
            f"Excel-отчёт отправлен пользователю {message.from_user.id}, записей: {len(orders)}"
        )

    except ImportError:
        await message.reply(
            "⚠️ Для генерации Excel-отчёта требуется библиотека openpyxl.\n"
            "Установите её: pip install openpyxl"
        )
    except Exception as e:
        logger.error(f"Ошибка при генерации Excel-отчёта: {e}", exc_info=True)
        await message.reply("⚠️ Произошла ошибка при генерации Excel-отчёта")


def _add_orders_sheet(workbook, worksheet, orders: list[dict], sheet_title: str) -> None:
    """Вспомогательная функция для добавления листа с заказами в Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    worksheet.title = sheet_title

    headers = [
        "ID",
        "Дата",
        "Менеджер",
        "Статус",
        "Заказ",
        "Цена",
        "Адрес",
        "Телефон",
        "Клиент",
        "Комментарий",
    ]

    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_num, order in enumerate(orders, 2):
        created_at = order.get("created_at", "") or ""
        if "T" in created_at:
            created_at = created_at.split("T")[0]

        values = [
            order.get("id", ""),
            created_at,
            order.get("manager_name", ""),
            order.get("status", ""),
            order.get("model", ""),
            order.get("price", ""),
            order.get("address", ""),
            order.get("phone", ""),
            order.get("customer_name", ""),
            order.get("comment", ""),
        ]

        for col_num, value in enumerate(values, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column].width = min(max_length + 2, 50)


async def build_daily_report_xlsx(for_date: date) -> bytes:
    """Генерирует ежедневный Excel-отчёт с 3 листами."""
    import openpyxl
    from openpyxl.styles import Font

    orders_today = await get_orders_for_date(for_date)
    all_orders = await get_orders_for_report()
    active_orders_today = await get_active_orders_for_date(for_date)
    
    logger.info(f"📊 Генерация ежедневного отчёта за {for_date}: всего заказов за день: {len(orders_today)}, активных: {len(active_orders_today)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Итоги за день")
    ws_active = wb.create_sheet("Заказы в работе")
    ws_all = wb.create_sheet("Все заказы")

    new_orders_count = len(orders_today)
    total_sum = 0
    status_counts = {}

    for order in orders_today:
        price_str = order.get("price", "") or "0"

        try:
            price = float(price_str)
        except (ValueError, TypeError):
            price = 0

        total_sum += price

        status = order.get("status", "")
        status_counts[status] = status_counts.get(status, 0) + 1

    ws_summary["A1"] = "Количество новых заказов за день:"
    ws_summary["B1"] = new_orders_count

    ws_summary["A2"] = "Общая сумма заказов за день:"
    ws_summary["B2"] = total_sum

    row = 3
    ws_summary[f"A{row}"] = "Распределение по статусам:"
    row += 1
    for status, count in status_counts.items():
        status_display = STATUS_DISPLAY.get(status, status)
        ws_summary[f"A{row}"] = f"  {status_display}:"
        ws_summary[f"B{row}"] = count
        row += 1

    for row in range(1, row):
        ws_summary[f"A{row}"].font = Font(bold=True)

    _add_orders_sheet(wb, ws_active, active_orders_today, "Заказы в работе")
    _add_orders_sheet(wb, ws_all, all_orders, "Все заказы")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    xlsx_bytes = buffer.read()
    buffer.close()

    return xlsx_bytes


@router.message(
    F.text.in_([BUTTON_DAILY_REPORT_TOGGLE_ON, BUTTON_DAILY_REPORT_TOGGLE_OFF])
)
async def handle_daily_report_toggle(message: Message) -> None:
    """Обработка кнопки переключения ежедневного отчёта."""
    if not await is_user_authorized(message.from_user.id):
        await message.reply("🔐 Доступ запрещён. Введите /start и PIN-код.")
        return
    try:
        chat_id = message.chat.id
        current_state = await get_daily_report_enabled(chat_id)

        new_state = not current_state
        await set_daily_report_enabled(chat_id, new_state)

        if new_state:
            await set_report_chat_id(chat_id, chat_id)
            await message.reply("✅ Ежедневный отчёт включен")
        else:
            await message.reply("❌ Ежедневный отчёт выключен")

        logger.info(
            f"Ежедневный отчёт для чата {chat_id} {'включен' if new_state else 'выключен'}"
        )

    except Exception as e:
        logger.error(f"Ошибка при переключении ежедневного отчёта: {e}", exc_info=True)
        await message.reply("⚠️ Произошла ошибка при изменении настройки")
